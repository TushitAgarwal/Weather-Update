__author__ = "TUSHIT AGARWAL"

from requests import get
import json
from openpyxl import Workbook, load_workbook
import os
from random import shuffle, choice
from time import sleep
import gzip
import shutil
import urllib
import urllib.request



def kelvin_to_Celcius(k):
    return (k - 273.15)


def Celcius_to_Fahrenheit(n): 
    return ((n * 9.0) / 5.0) + 32.0


def createExcelSheet():

    wb = Workbook()

    wb.create_sheet(title = "Weather", index = 0)
    wb.create_sheet(title = "City Tokens", index = 1)
    wb.create_sheet(title = "Additional Details", index = 2)
    

    sheet = wb['City Tokens'] 

    sheet.append(("CITY", "CITY_ID"))

    temp = []

    with open(os.path.join(os.getcwd(), "city.list.json"), 'r', encoding = 'utf-8') as f:
        for i in json.loads(f.read()):
           sheet.append((i['name'], i['id']))
           temp.append((i['name'], i['id']))
    
    sheet = wb["Weather"]

    sheet.append(["CITY TOKEN", "TEMPERATURE", "HUMIDITY", "UNIT", "STATE UPDATE(0/1)"])


    for i in range(20):
        sheet.append([temp[i][1], "", "", choice(["C", "F"]), choice([0,1])])

    wb.save('task1.xlsx') 


def wait(n):
    c = 0
    t = ["/", "-", "\\", "|"]
    while c <= n:
        sleep(1)
        print(f"Waiting for {n} seconds...{t[c%4]}", end = "\r")
        c += 1


def main():
    if os.path.isfile(os.path.join(os.getcwd(), "task1.xlsx")):
        try:
            print("Loading Workbook...", end = "\r")
            wb = load_workbook("task1.xlsx")
        
        except:
            print("Failed to load Workbook due to some error. Deleting Existing Workbook...", end = "\r")
            sleep(2)
            os.remove("task1.xlsx")
            print("Creating Workbook..." + " " * 50, end = "\r")
            createExcelSheet()
        
        finally:
            wb = load_workbook("task1.xlsx")
            print("Workbook loaded successfully..." + " " * 50, end = "\r")

    else:
        print("No Workbook found... Creating a new Workbook" + " " * 50, end = "\r")
        createExcelSheet()
        wb = load_workbook("task1.xlsx")
    
    sheet = wb["Weather"]

    while True:
        print("Fetching API..." + " " * 30, end = "\r")
        try:
            for row in sheet.iter_rows():
                cityToken, temperature, humidity, unit, update = map(lambda x: x.value, row)
                
                if update == 1:
                    d = json.loads(get(f"http://api.openweathermap.org/data/2.5/weather?id={cityToken}&appid=5a4581a471721bb8b3e1e2df99814958").text)
                    if unit == "F":
                        temperature = "%.3f" %kelvin_to_Celcius(Celcius_to_Fahrenheit(d["main"]["temp"]))
                    elif unit == "C":
                        temperature = "%.3f" %kelvin_to_Celcius(d["main"]["temp"])
                    humidity = d["main"]["humidity"]
                    data = [cityToken, temperature, humidity, unit, update]
                    for i in range(5):
                        row[i].value = data[i]
            
            print("Updating Values to Workbook and saving file....", end = "\r")
            wb.save("task1.xlsx")
            print(" " * 100, end = "\r") # For clearing the output stream.
            wait(10)

        except KeyboardInterrupt:
            print("Exiting program...Saving the file..." + " " * 50, end = "\r")

            wb.save("task1.xlsx")
            break
        except:
            print("Unknown error occurs...exiting the program..saving the file...", end = "\r")
            wb.save("task1.xlsx")
            break
        
    print('Workbook updated successfully...' + " " * 50)



if __name__ == "__main__":
    opener = urllib.request.URLopener()
    opener.retrieve("http://bulk.openweathermap.org/sample/city.list.json.gz", "city.list.json.gz")
    with gzip.open('city.list.json.gz', 'rb') as f_in:
        with open('city.list.json', 'wb') as f_out:
            shutil.copyfileobj(f_in, f_out)
    main()
